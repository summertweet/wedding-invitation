#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
婚礼请柬 — 宾客链接批量生成脚本
用法:
  python3 generate_links.py                    # 从 guests.csv 读取，输出到终端
  python3 generate_links.py --output links.csv # 输出到 CSV 文件
  python3 generate_links.py --output links.xlsx # 输出到 Excel 文件
  python3 generate_links.py --copy              # 复制全部链接到剪贴板
"""

import csv
import sys
import os
import urllib.parse
import subprocess
import argparse

# ==================== 配置 ====================
# 👇 部署后修改为你的实际网址
BASE_URL = "https://summertweet.github.io/wedding-invitation/"

# 宾客名单文件 (CSV 格式)
GUEST_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "guests.csv")
# ================================================


def read_guests(filepath):
    """读取宾客名单"""
    guests = []
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get('姓名', '').strip()
            if name:
                guests.append({
                    'name': name,
                    'title': row.get('称呼', '').strip(),
                    'note': row.get('备注', '').strip(),
                })
    return guests


def generate_link(base_url, guest_name):
    """生成个人专属链接"""
    encoded = urllib.parse.quote(guest_name)
    url = base_url.rstrip('/') + f'/?name={encoded}'
    return url


def output_terminal(guests, base_url):
    """输出到终端"""
    print("\n" + "=" * 60)
    print(f"  💍 婚礼请柬链接  (共 {len(guests)} 位宾客)")
    print("=" * 60)
    print(f"  基础网址: {base_url}")
    print("-" * 60)

    for i, g in enumerate(guests, 1):
        link = generate_link(base_url, g['name'])
        note = f"  ({g['note']})" if g['note'] else ""
        print(f"\n  [{i:>3}] {g['name']}{note}")
        print(f"        {link}")

    print("\n" + "=" * 60)
    print(f"  ✅ 共生成 {len(guests)} 个链接")
    print("=" * 60 + "\n")


def output_csv(guests, base_url, output_path):
    """输出到 CSV 文件"""
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['序号', '姓名', '称呼', '备注', '专属链接'])
        for i, g in enumerate(guests, 1):
            link = generate_link(base_url, g['name'])
            writer.writerow([i, g['name'], g['title'], g['note'], link])

    print(f"\n✅ 已生成 CSV 文件: {output_path}")
    print(f"   共 {len(guests)} 位宾客")


def output_excel(guests, base_url, output_path):
    """输出到 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("⚠️  需要安装 openpyxl: pip3 install openpyxl")
        print("   正在自动安装...")
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], check=True)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "宾客链接"

    # 样式定义
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='D4797A', end_color='D4797A', fill_type='solid')
    cell_font = Font(name='微软雅黑', size=11)
    link_font = Font(name='Consolas', size=10, color='0563C1', underline='single')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    # 表头
    headers = ['序号', '姓名', '称呼', '备注', '专属链接']
    col_widths = [8, 15, 15, 20, 60]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 数据行
    for i, g in enumerate(guests, 1):
        link = generate_link(base_url, g['name'])
        row = i + 1

        ws.cell(row=row, column=1, value=i).font = cell_font
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=g['name']).font = Font(name='微软雅黑', size=11, bold=True)
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=g['title']).font = cell_font
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=g['note']).font = cell_font
        ws.cell(row=row, column=4).alignment = left
        ws.cell(row=row, column=4).border = thin_border

        cell = ws.cell(row=row, column=5, value=link)
        cell.font = link_font
        cell.hyperlink = link
        cell.alignment = left
        cell.border = thin_border

        # 交替行背景色
        if i % 2 == 0:
            row_fill = PatternFill(start_color='FDF0F0', end_color='FDF0F0', fill_type='solid')
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = row_fill

    # 冻结表头
    ws.freeze_panes = 'A2'

    # 添加统计信息行
    stats_row = len(guests) + 3
    ws.cell(row=stats_row, column=1, value=f'共 {len(guests)} 位宾客').font = Font(name='微软雅黑', size=10, color='888888')
    ws.cell(row=stats_row + 1, column=1, value=f'基础网址: {base_url}').font = Font(name='微软雅黑', size=10, color='888888')

    wb.save(output_path)
    print(f"\n✅ 已生成 Excel 文件: {output_path}")
    print(f"   共 {len(guests)} 位宾客")
    print(f"   可直接用 Excel/WPS 打开，点击链接即可访问")


def copy_to_clipboard(guests, base_url):
    """复制全部链接到剪贴板"""
    lines = []
    for g in guests:
        link = generate_link(base_url, g['name'])
        lines.append(f"{g['name']}: {link}")
    text = '\n'.join(lines)

    try:
        process = subprocess.Popen(['pbcopy'], stdin=subprocess.PIPE)
        process.communicate(text.encode('utf-8'))
        print(f"\n✅ 已复制 {len(guests)} 个链接到剪贴板")
    except FileNotFoundError:
        try:
            process = subprocess.Popen(['xclip', '-selection', 'clipboard'], stdin=subprocess.PIPE)
            process.communicate(text.encode('utf-8'))
            print(f"\n✅ 已复制 {len(guests)} 个链接到剪贴板")
        except FileNotFoundError:
            print("⚠️ 无法复制到剪贴板，请手动复制以下内容：")
            print(text)


def main():
    parser = argparse.ArgumentParser(description='婚礼请柬宾客链接批量生成')
    parser.add_argument('--input', '-i', default=GUEST_FILE, help='宾客名单文件路径 (CSV)')
    parser.add_argument('--output', '-o', help='输出文件路径 (.csv 或 .xlsx)')
    parser.add_argument('--copy', '-c', action='store_true', help='复制链接到剪贴板')
    parser.add_argument('--url', '-u', default=BASE_URL, help='请柬基础网址')
    args = parser.parse_args()

    # 读取宾客
    if not os.path.exists(args.input):
        print(f"❌ 找不到宾客名单文件: {args.input}")
        print(f"   请创建 CSV 文件，格式：姓名,称呼,备注")
        sys.exit(1)

    guests = read_guests(args.input)
    if not guests:
        print("❌ 宾客名单为空")
        sys.exit(1)

    print(f"\n📋 读取到 {len(guests)} 位宾客")

    # 输出
    if args.output:
        ext = os.path.splitext(args.output)[1].lower()
        if ext == '.xlsx':
            output_excel(guests, args.url, args.output)
        else:
            output_csv(guests, args.url, args.output)
    elif args.copy:
        copy_to_clipboard(guests, args.url)
    else:
        output_terminal(guests, args.url)

    # 始终输出到终端
    if args.output or args.copy:
        output_terminal(guests, args.url)


if __name__ == '__main__':
    main()
